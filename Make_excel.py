import calendar
from datetime import date

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

# 액셀 스타일 함수 (export_excel 함수내부 사용)
def _style_excel(scheduler, output_path: str) -> None:
    wb = load_workbook(output_path)

    ws = wb["모니터링표"]
    comp_ws = wb["대체휴무"]
    stats_ws = wb["통계"]

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    weekend_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    holiday_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    comp_off_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center_alignment = Alignment(horizontal="center", vertical="center")

    for target_ws in [ws, comp_ws, stats_ws]:
        for cell in target_ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border

    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    day_type_col = headers["구분"]
    comp_off_col = headers["대체휴무일"]

    for row in ws.iter_rows(min_row=2):
        day_type_value = row[day_type_col - 1].value
        comp_off_value = row[comp_off_col - 1].value

        for cell in row:
            cell.border = border
            cell.alignment = center_alignment

        if day_type_value in ("토요일", "일요일"):
            for cell in row:
                cell.fill = weekend_fill

        if day_type_value == "공휴일":
            for cell in row:
                cell.fill = holiday_fill

        if comp_off_value:
            row[comp_off_col - 1].fill = comp_off_fill

    for row in comp_ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = center_alignment
        if row[0].value:
            for cell in row:
                cell.fill = comp_off_fill

    for row in stats_ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = center_alignment

    _add_unavailable_table_to_sheet(scheduler, ws)

    for target_ws in [ws, comp_ws, stats_ws]:
        for col_idx, column_cells in enumerate(target_ws.columns, start=1):
            max_length = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            target_ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 3

        target_ws.freeze_panes = "A2"

    wb.save(output_path)

# 1. 캘린더 시트 추가 함수 (export_excel 함수내부 사용)
# 개인별 모니터링 일정 캘린더
def _add_calendar_sheet(scheduler, output_path: str, df: pd.DataFrame) -> None:
    wb = load_workbook(output_path)

    if "캘린더" in wb.sheetnames:
        del wb["캘린더"]

    ws = wb.create_sheet("캘린더", 0)

    year = scheduler.config.year
    month = scheduler.config.month

    ws.merge_cells("A1:G1")
    ws["A1"] = f"{year}년 {month}월 모니터링표"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    weekdays = ["일", "월", "화", "수", "목", "금", "토"]

    for col_idx, weekday in enumerate(weekdays, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = weekday
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    df_by_date = {
        row["날짜"]: row
        for _, row in df.iterrows()
    }

    cal = calendar.Calendar(firstweekday=6).monthdayscalendar(year, month)

    start_row = 3

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sunday_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    saturday_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    holiday_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    normal_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    empty_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")

    for week_idx, week in enumerate(cal):
        row_idx = start_row + week_idx

        for col_idx, day in enumerate(week, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)

            cell.border = border
            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )

            if day == 0:
                cell.value = ""
                cell.fill = empty_fill
                continue

            current_date = date(year, month, day)
            data = df_by_date.get(current_date)

            if data is None:
                continue

            day_type = data["구분"]
            assignee = data["모니터링 담당자"]
            comp_off = data["대체휴무일"]

            comp_off_text = ""
            if pd.notna(comp_off):
                if hasattr(comp_off, "strftime"):
                    comp_off_text = comp_off.strftime("%Y-%m-%d")
                else:
                    comp_off_text = str(comp_off)

            rich_text = CellRichText()

            rich_text.append(f"{day}일 ")
            rich_text.append(TextBlock(InlineFont(b=True), f"[{day_type}]"))
            rich_text.append("\n")

            rich_text.append("담당자: ")
            rich_text.append(TextBlock(InlineFont(b=True), str(assignee)))

            if comp_off_text:
                rich_text.append("\n대휴: ")
                rich_text.append(TextBlock(InlineFont(u="single"), comp_off_text))

            cell.value = rich_text

            if day_type == "공휴일":
                cell.fill = holiday_fill
            elif day_type == "일요일":
                cell.fill = sunday_fill
            elif day_type == "토요일":
                cell.fill = saturday_fill
            else:
                cell.fill = normal_fill

    for col_idx in range(1, 8):
        ws.column_dimensions[get_column_letter(col_idx)].width = 24

    for row_idx in range(3, 9):
        ws.row_dimensions[row_idx].height = 95

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24

    wb.save(output_path)

# 2. 모니터링표 시트 추가 함수 (export_excel 함수내부 사용)
# 개인별 불가능 일정 표 추가
def _add_unavailable_table_to_sheet(scheduler, ws) -> None:
    start_col = ws.max_column + 2
    start_row = 1

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    name_col = start_col
    date_col = start_col + 1

    name_header = ws.cell(row=start_row, column=name_col)
    date_header = ws.cell(row=start_row, column=date_col)

    name_header.value = "이름"
    date_header.value = "불가능 일정"

    for cell in [name_header, date_header]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    current_row = start_row + 1

    for member in scheduler.all_members:
        unavailable_dates = sorted(
            scheduler.config.personal_unavailable.get(member, set())
        )

        if not unavailable_dates:
            continue

        start_merge_row = current_row

        for unavailable_date in unavailable_dates:
            date_text = (
                f"{unavailable_date.strftime('%Y-%m-%d')} "
                f"({['월', '화', '수', '목', '금', '토', '일'][unavailable_date.weekday()]})"
            )

            ws.cell(row=current_row, column=date_col).value = date_text
            ws.cell(row=current_row, column=date_col).alignment = left
            ws.cell(row=current_row, column=date_col).border = border

            ws.cell(row=current_row, column=name_col).border = border

            current_row += 1

        end_merge_row = current_row - 1

        if start_merge_row <= end_merge_row:
            ws.merge_cells(
                start_row=start_merge_row,
                start_column=name_col,
                end_row=end_merge_row,
                end_column=name_col,
            )

            name_cell = ws.cell(row=start_merge_row, column=name_col)
            name_cell.value = member
            name_cell.alignment = center
            name_cell.border = border

            for row_idx in range(start_merge_row, end_merge_row + 1):
                ws.cell(row=row_idx, column=name_col).border = border

    ws.column_dimensions[get_column_letter(name_col)].width = 14
    ws.column_dimensions[get_column_letter(date_col)].width = 28

# 3. 대체휴무 시트 생성 함수 (export_excel 함수내부 사용)
# 근무일 유형에 따른 개인별 대체휴무일
def _comp_stats(scheduler, df: pd.DataFrame) -> pd.DataFrame:
    comp_rows = []

    for _, row in df.iterrows():
        if pd.notna(row["대체휴무일"]):
            comp_rows.append({
                "대체휴무일": row["대체휴무일"],
                "담당자": row["모니터링 담당자"],
                "근무일": row["날짜"],
                "근무구분": row["구분"],
            })

    comp_df = pd.DataFrame(comp_rows)

    if not comp_df.empty:
        comp_df["대체휴무일"] = pd.to_datetime(comp_df["대체휴무일"]).dt.date
        comp_df = comp_df.sort_values(
            by=["대체휴무일"],
            ascending=True,
        )

    return comp_df

# 4. 통계 시트 추가 함수 (export_excel 함수내부 사용)
# 담당자 별 모니터링 횟수 통계
def _build_stats(scheduler, df: pd.DataFrame) -> pd.DataFrame:
    stats_rows = []

    for member in scheduler.all_members:
        member_df = df[df["모니터링 담당자"] == member]

        weekday_count = (member_df["구분"] == "평일").sum()
        saturday_count = (member_df["구분"] == "토요일").sum()
        sunday_count = (member_df["구분"] == "일요일").sum()
        holiday_count = (member_df["구분"] == "공휴일").sum()

        stats_rows.append({
            "이름": member,
            "평일 모니터링횟수": int(weekday_count),
            "주말 모니터링횟수": int(saturday_count + sunday_count),
            "공휴일 모니터링횟수": int(holiday_count),
            "총 모니터링횟수": int(len(member_df)),
        })

    stats_df = pd.DataFrame(stats_rows).sort_values(
        by=["총 모니터링횟수", "이름"],
        ascending=[False, True],
    )

    return stats_df

# 액셀에서 사용할 Config 기준 데이터 값 불러오는 함수
def export_excel(scheduler, output_path: str, df: pd.DataFrame) -> pd.DataFrame:
    df = df.rename(columns={
        "date": "날짜",
        "weekday_name": "요일",
        "day_type": "구분",
        "rotation_type": "로테이션",
        "assignee": "모니터링 담당자",
        "comp_off_date": "대체휴무일",
    })

    comp_df = _comp_stats(scheduler, df)
    stats_df = _build_stats(scheduler, df)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="모니터링표", index=False)
        comp_df.to_excel(writer, sheet_name="대체휴무", index=False)
        stats_df.to_excel(writer, sheet_name="통계", index=False)

    _add_calendar_sheet(scheduler, output_path, df)
    _style_excel(scheduler, output_path)

    return df
